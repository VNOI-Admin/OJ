import fnmatch
import json
import openpyxl
import os
import re
import secrets
import tempfile
import zipfile

from celery import shared_task
from django.conf import settings
from django.contrib.auth.models import User
from django.core.cache import cache
from django.urls import reverse
from django.utils.translation import gettext as _

from judge.models import Comment, Language, Problem, Profile, Submission
from judge.models.profile import Organization
from judge.utils.celery import Progress
from judge.utils.raw_sql import use_straight_join
from judge.utils.unicode import utf8bytes

__all__ = ('prepare_user_data', 'bulk_create_users')
rewildcard = re.compile(r'\*+')


def apply_submission_filter(queryset, options):
    if not options['submission_download']:
        return []

    use_straight_join(queryset)

    if options['submission_results']:
        queryset = queryset.filter(result__in=options['submission_results'])

    # Compress wildcards to avoid exponential complexity on certain glob patterns before Python 3.9.
    # For details, see <https://bugs.python.org/issue40480>.
    problem_glob = rewildcard.sub('*', options['submission_problem_glob'])
    if problem_glob != '*':
        queryset = queryset.filter(
            problem__in=Problem.objects.filter(code__regex=fnmatch.translate(problem_glob)),
        )

    return list(queryset)


def apply_comment_filter(queryset, options):
    if not options['comment_download']:
        return []
    return list(queryset)


@shared_task(bind=True)
def prepare_user_data(self, profile_id, options):
    options = json.loads(options)
    with Progress(self, 2, stage=_('Applying filters')) as p:
        # Force an update so that we get a progress bar.
        p.done = 0
        submissions = apply_submission_filter(
            Submission.objects.select_related('problem', 'language', 'source').filter(user_id=profile_id),
            options,
        )
        p.did(1)
        comments = apply_comment_filter(Comment.objects.filter(author_id=profile_id), options)
        p.did(1)

    with zipfile.ZipFile(os.path.join(settings.DMOJ_USER_DATA_CACHE, '%s.zip' % profile_id), mode='w') as data_file:
        submission_count = len(submissions)
        if submission_count:
            submission_info = {}
            with Progress(self, submission_count, stage=_('Preparing your submission data')) as p:
                prepared = 0
                interval = max(submission_count // 10, 1)
                for submission in submissions:
                    submission_info[submission.id] = {
                        'problem': submission.problem.code,
                        'date': submission.date.isoformat(),
                        'time': submission.time,
                        'memory': submission.memory,
                        'language': submission.language.key,
                        'status': submission.status,
                        'result': submission.result,
                        'case_points': submission.case_points,
                        'case_total': submission.case_total,
                    }
                    with data_file.open(
                        'submissions/%s.%s' % (submission.id, submission.language.extension),
                        'w',
                    ) as f:
                        f.write(utf8bytes(submission.source.source))

                    prepared += 1
                    if prepared % interval == 0:
                        p.done = prepared

                with data_file.open('submissions/info.json', 'w') as f:
                    f.write(utf8bytes(json.dumps(submission_info, sort_keys=True, indent=4)))

        comment_count = len(comments)
        if comment_count:
            comment_info = {}
            with Progress(self, comment_count, stage=_('Preparing your comment data')) as p:
                prepared = 0
                interval = max(comment_count // 10, 1)
                for comment in comments:
                    related_object = {
                        'b': 'blog post',
                        'c': 'contest',
                        'p': 'problem',
                        's': 'problem editorial',
                    }
                    comment_info[comment.id] = {
                        'date': comment.time.isoformat(),
                        'related_object': related_object[comment.page[0]],
                        'page': comment.page[2:],
                        'score': comment.score,
                    }
                    with data_file.open('comments/%s.txt' % comment.id, 'w') as f:
                        f.write(utf8bytes(comment.body))

                    prepared += 1
                    if prepared % interval == 0:
                        p.done = prepared

                with data_file.open('comments/info.json', 'w') as f:
                    f.write(utf8bytes(json.dumps(comment_info, sort_keys=True, indent=4)))

    return submission_count + comment_count


ALPHABET = 'abcdefghkqtxyz' + 'abcdefghkqtxyz'.upper() + '23456789'


def generate_password():
    return ''.join(secrets.choice(ALPHABET) for _ in range(8))


def validate_user_data(users_data):
    """Validate user data and check for existing users"""
    errors = []
    seen_usernames = set()

    # Check for existing users in database
    usernames = [user['username'] for user in users_data]

    existing_users = User.objects.filter(username__in=usernames)
    if existing_users.exists():
       return _('Some usernames already exist in the database')

    # Validate each user
    for i, user in enumerate(users_data):
        row_num = i + 2  # Excel rows start from 1, plus header row
        username = user.get('username', '').strip()
        name = user.get('name', '').strip()

        # Check required fields
        if not username:
            return _('Row %d: Username is required') % row_num
        if not name:
            return _('Row %d: Name is required') % row_num

        if not username or not name:
            continue

        # Check username format
        if not re.match(r'^[a-zA-Z0-9_.-]+$', username):
            return _('Row %d: Username can only contain letters, numbers, dots, hyphens, and underscores') % row_num
        # cannot be > 20 length
        if len(username) > 20:
            return _('Row %d: Username cannot be longer than 20 characters') % row_num

        # Check for duplicates in file
        if username in seen_usernames:
            return _('Row %d: Username "%s" appears multiple times in the file') % (row_num, username)
        else:
            seen_usernames.add(username)

    return None


@shared_task(bind=True)
def bulk_create_users(self, file_path, admin_user_id, org_pk_to_join):
    """Create users in bulk from Excel file"""

    try:
        # Read Excel file using openpyxl
        workbook = openpyxl.load_workbook(file_path)
        worksheet = workbook.active

        # Get headers from first row
        headers = []
        for cell in worksheet[1]:
            if cell.value:
                headers.append(cell.value.lower().strip())
            else:
                headers.append('')

        # Check required columns
        required_columns = ['username', 'name']
        missing_columns = [col for col in required_columns if col not in headers]
        if missing_columns:
            raise Exception(_('Missing required columns: %s') % ', '.join(missing_columns))

        # Get column indices for required columns
        column_indices = {}
        for i, header in enumerate(headers):
            if header in required_columns:
                column_indices[header] = i

        # Read data rows
        users_data = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(row):  # Skip empty rows
                continue

            user_data = {}
            for col_name, col_index in column_indices.items():
                value = row[col_index] if col_index < len(row) and row[col_index] is not None else ''
                user_data[col_name] = str(value).strip()

            # Only add row if it has at least one non-empty value
            if any(user_data.values()):
                users_data.append(user_data)

        # Validate data
        with Progress(self, 1, stage=_('Validating user data')) as p:
            errors = validate_user_data(users_data)
            p.did(1)

        if errors:
            # Store errors in cache for retrieval
            cache.set(f'bulk_user_errors_{self.request.id}', [errors], 3600)  # 1 hour
            raise Exception(_('Validation errors found'))

        # Create users
        created_users = []
        total_users = len(users_data)

        org_to_join = None
        if org_pk_to_join is not None:
            try:
                org_to_join = Organization.objects.get(pk=org_pk_to_join)
            except Organization.DoesNotExist:
                org_to_join = None

        with Progress(self, total_users, stage=_('Creating user accounts')) as p:
            default_language = Language.objects.get(key=settings.DEFAULT_USER_LANGUAGE)

            for i, user_data in enumerate(users_data):
                username = user_data['username']
                name = user_data['name']
                password = generate_password()

                # Create Django user
                user = User(username=username, first_name=name, is_active=True)
                user.set_password(password)
                user.save()

                # Create profile
                profile = Profile(user=user)
                profile.language = default_language
                profile.username_display_override = name
                profile.save()

                if org_to_join is not None:
                    profile.organizations.add(org_to_join)

                created_users.append({
                    'username': username,
                    'name': name,
                    'password': password,
                })

                p.did(1)

        # Generate Excel output
        from openpyxl import Workbook

        output_path = os.path.join(
            settings.DMOJ_USER_DATA_CACHE or tempfile.gettempdir(),
            f'bulk_users_{self.request.id}.xlsx'
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Users"

        # Write header
        fieldnames = ['username', 'name', 'password']
        ws.append(fieldnames)

        # Write user data
        for user in created_users:
            ws.append([user['username'], user['name'], user['password']])

        # Save Excel file
        wb.save(output_path)
        print(output_path)

        # Clean up input file
        try:
            os.remove(file_path)
        except OSError:
            pass

        return {
            'success': True,
            'created_count': len(created_users),
            'output_file': output_path,
        }

    except Exception as e:
        # Clean up input file on error
        try:
            os.remove(file_path)
        except OSError:
            pass

        # If validation errors, re-raise with original message
        if 'Validation errors found' in str(e):
            raise e

        # For other errors, wrap in a more user-friendly message
        raise Exception(_('Error processing file: %s') % str(e))
